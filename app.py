
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
from io import BytesIO

st.set_page_config(page_title="Planning GP (22/09 → 04/12)", layout="wide")

st.title("📅 Générateur d'emplois du temps par GP (22/09/2025 → 04/12/2025)")
st.caption("Saisie/édition des créneaux, aperçu, export Excel multi-feuilles (une feuille par GP). Jours en français, créneaux : 8h30–10h30 / 10h45–12h45 / 13h45–15h45 / 16h–18h.")

# ---------- Données initiales ----------
RAW = [
("24/09/2025","Biochimie TD1 GP6","13h45","15h45"),
("24/09/2025","Biochimie TD1 GP5","16h","18h"),
("25/09/2025","Biochimie TD1 GP4","8h30","10h30"),
("25/09/2025","Biochimie TD1 GP3","10h45","12h45"),
("25/09/2025","Biochimie TD1 GP1","13h45","15h45"),
("25/09/2025","Biochimie TD1 GP2","16h","18h"),
("01/10/2025","Biochimie TD1 GP12","13h45","15h45"),
("01/10/2025","Biochimie TD1 GP11","16h","18h"),
("02/10/2025","Biochimie TD1 GP10","8h30","10h30"),
("02/10/2025","Biochimie TD1 GP9","10h45","12h45"),
("02/10/2025","Biochimie TD1 GP8","13h45","15h45"),
("02/10/2025","Biochimie TD1 GP7","16h","18h"),
("08/10/2025","Immunologie TD1 G2","13h45","15h45"),
("08/10/2025","Biologie moléculaire TD1 G1","13h45","15h45"),
("08/10/2025","Immunologie TD1 G1","16h","18h"),
("08/10/2025","Biologie moléculaire TD1 G2","16h","18h"),
("09/10/2025","Immunologie TD1 G3","8h30","10h30"),
("09/10/2025","Biologie moléculaire TD1 G4","8h30","10h30"),
("09/10/2025","Immunologie TD1 G4","10h45","12h45"),
("09/10/2025","Biologie Moléculaire TD1 G3","10h45","12h45"),
("09/10/2025","Immunologie TD1 G6","13h45","15h45"),
("09/10/2025","Biologie moléculaire TD1 G5","13h45","15h45"),
("09/10/2025","Immunologie TD1 G5","16h","18h"),
("09/10/2025","Biologie moléculaire TD1 G6","16h","18h"),
("15/10/2025","Biochimie TD2 GP6","13h45","15h45"),
("15/10/2025","Immunologie TD1 GP7","13h45","15h45"),
("15/10/2025","Biologie moléculaire TD1 GP8","13h45","15h45"),
("15/10/2025","Biochimie TD2 GP5","16h","18h"),
("15/10/2025","Immunologie TD1 G8","16h","18h"),
("15/10/2025","Biologie moléculaire TD1 G7","16h","18h"),
("16/10/2025","Biochimie TD2 GP4","8h30","10h30"),
("16/10/2025","Immunologie TD1 GP10","8h30","10h30"),
("16/10/2025","Biologie moléculaire TD1 GP9","8h30","10h30"),
("16/10/2025","Biochimie TD2 GP3","10h45","12h45"),
("16/10/2025","Immunologie TD1 GP9","10h45","12h45"),
("16/10/2025","Biologie moléculaire TD1 G10","10h45","12h45"),
("16/10/2025","Biochimie TD2 GP2","13h45","15h45"),
("16/10/2025","Immunologie TD1 G11","13h45","15h45"),
("16/10/2025","Biologie moléculaire TD1 G12","13h45","15h45"),
("16/10/2025","Biochimie TD2 GP1","16h","18h"),
("16/10/2025","Immunologie TD1 G12","16h","18h"),
("16/10/2025","Biologie moléculaire TD1 G11","16h","18h"),
("22/10/2025","Biochimie TD2 GP12","13h45","15h45"),
("22/10/2025","Biochimie TD2 GP11","16h","18h"),
("22/10/2025","Biochimie TD2 GP10","8h30","10h30"),
("22/10/2025","Biochimie TD2 GP9","10h45","12h45"),
("22/10/2025","Biochimie TD2 GP8","13h45","15h45"),
("22/10/2025","Biochimie TD2 GP7","16h","18h"),
("29/10/2025","Immunologie TD2 GP2","13h45","15h45"),
("29/10/2025","Biologie moléculaire TD2 G1","13h45","15h45"),
("29/10/2025","Immunologie TD2 GP1","16h","18h"),
("29/10/2025","Biologie moléculaire TD2 G2","16h","18h"),
("30/10/2025","Immunologie TD2 G3","8h30","10h30"),
("30/10/2025","Biologie moléculaire TD2 G4","8h30","10h30"),
("30/10/2025","Immunologie TD2 GP4","10h45","12h45"),
("30/10/2025","Biologie molécularie TD2 GP3","10h45","12h45"),
("30/10/2025","Immunologie TD2 GP5","13h45","15h45"),
("30/10/2025","Biologie moléculaire TD2 GP6","13h45","15h45"),
("30/10/2025","Immunologie TD2 GP6","16h","18h"),
("30/10/2025","Biologie moléculaire TD2 G5","16h","18h"),
("05/11/2025","Biochimie TD3 GP1","13h45","15h45"),
("05/11/2025","Communication cellulaire et signalisation TD1 GP2","13h45","15h45"),
("05/11/2025","Immunologie TD2 GP7","13h45","15h45"),
("05/11/2025","Biologie moléculaire TD2 GP8","13h45","15h45"),
("05/11/2025","Biochimie TD3 GP2","16h","18h"),
("05/11/2025","Communication cellulaire et signalisation TD1 GP1","16h","18h"),
("05/11/2025","Immunologie TD2 GP8","16h","18h"),
("05/11/2025","Biologie moléculaire TD2 GP7","16h","18h"),
("06/11/2025","Biochimie TD3 GP3","8h30","10h30"),
("06/11/2025","Communication cellulaire et signalisation TD1 GP4","8h30","10h30"),
("06/11/2025","Immunologie TD2 GP10","8h30","10h30"),
("06/11/2025","Biologie moléculaire TD2 GP9","8h30","10h30"),
("06/11/2025","Biochimie TD3 GP4","10h45","12h45"),
("06/11/2025","Communication cellulaire et signalisation TD1 GP3","10h45","12h45"),
("06/11/2025","Immunologie TD2 GP9","10h45","12h45"),
("06/11/2025","Biologie moléculaire TD2 GP10","10h45","12h45"),
("06/11/2025","Biochimie TD3 GP5","13h45","15h45"),
("06/11/2025","Communication cellulaire et signalisation TD1 GP6","13h45","15h45"),
("06/11/2025","Immunologie TD2 GP11","13h45","15h45"),
("06/11/2025","biologie moléculaire td2 GP12","13h45","15h45"),
("06/11/2025","Biochimie TD3 GP6","16h","18h"),
("06/11/2025","Communication cellulaire et signalisation TD1 G5","16h","18h"),
("06/11/2025","Immunologie TD2 GP12","16h","18h"),
("06/11/2025","Biologie moléculaire TD2 GP11","16h","18h"),
("12/11/2025","Biochimie TD3 GP7","13h45","15h45"),
("12/11/2025","Communication cellulaire et signalisation GP8","13h45","15h45"),
("12/11/2025","Biochimie TD3 GP8","16h","18h"),
("12/11/2025","Communication cellulaire et signalisation TD1 GP7","16h","18h"),
("13/11/2025","Biochimie TD3 GP11","8h30","10h30"),
("13/11/2025","Communication cellulaire et signalisation TD1 GP12","8h30","10h30"),
("13/11/2025","Biochimie TD3 GP12","10h45","12h45"),
("13/11/2025","Communication cellulaire et signalisation TD1 GP11","10h45","12h45"),
("13/11/2025","Biochimie TD3 GP9","13h45","15h45"),
("13/11/2025","Communication cellulaire et signalisation TD1 GP10","13h45","15h45"),
("13/11/2025","Biochimie TD3 GP10","16h","18h"),
("13/11/2025","Communication cellulaire et signalisation TD1 GP9","16h","18h"),
("19/11/2025","Immunologie TD3 G2","13h45","15h45"),
("19/11/2025","Biologie moléculaire TD3 G1","13h45","15h45"),
("19/11/2025","Immunologie TD3 GP1","16h","18h"),
("19/11/2025","Biologie moléculaire TD3 GP2","16h","18h"),
("20/11/2025","Immunologie TD3 GP6","8h30","10h30"),
("20/11/2025","Biologie moléculaire TD3 GP7","8h30","10h30"),
("20/11/2025","Immunologie TD3 GP7","10h30","12h45"),
("20/11/2025","Biologie moléculaire TD3 GP6","10h45","12h45"),
("20/11/2025","Immunologie TD3 G4","13h45","15h45"),
("20/11/2025","Biologie moléculaire TD3 GP3","13h45","15h45"),
("20/11/2025","Immunologie TD3 GP3","16h","18h"),
("20/11/2025","Biologie moléculaire TD3 GP4","16h","18h"),
("26/11/2025","Biochimie TD4 GP1","13h45","15h45"),
("26/11/2025","Communication cellulaire et signalisation TD2 GP2","13h45","15h45"),
("26/11/2025","Immunologie TD3 GP5","13h45","15h45"),
("26/11/2025","Biologie moléculaire TD3","13h45","15h45"),
("26/11/2025","Biochimie TD4 GP2","16h","18h"),
("26/11/2025","Communication cellulaire et signalisation TD2 GP1","16h","18h"),
("26/11/2025","Immunologie TD3 GP8","16h","18h"),
("26/11/2025","Biologie moléculaire TD3 GP7","16h","18h"),
("27/11/2025","Biochimie TD4 GP3","8h30","10h30"),
("27/11/2025","Communication cellulaire et signalisation TD2 GP4","8h30","10h30"),
("27/11/2025","Immunologie TD3 GP10","8h30","10h30"),
("27/11/2025","Biologie moléculaire TD3 GP9","8h30","10h30"),
("27/11/2025","Biochimie TD4 GP4","10h45","12h45"),
("27/11/2025","Communication cellulaire et signalisation TD2 GP3","10h45","12h45"),
("27/11/2025","Immunologie TD3 GP9","10h45","12h45"),
("27/11/2025","Biologie moléculaire TD3 GP10","10h45","12h45"),
("27/11/2025","Biochimie TD4 GP5","13h45","15h45"),
("27/11/2025","Communication cellulaire et signalisation TD2 GP6","13h45","15h45"),
("27/11/2025","Immunologie TD3 GP11","13h45","15h45"),
("27/11/2025","Biologie moléculaire TD3 GP12","13h45","15h45"),
("27/11/2025","Biochimie TD4 GP6","16h","18h"),
("27/11/2025","Communication cellulaire et signalisation TD2 GP5","16h","18h"),
("27/11/2025","Immunologie TD3 GP12","16h","18h"),
("27/11/2025","Biologie moléculaire TD3 GP11","16h","18h"),
("03/12/2025","Biochimie TD4 GP7","13h45","15h45"),
("03/12/2025","Communication cellulaire et signalisation TD2 GP8","13h45","15h45"),
("03/12/2025","Biochimie TD4 GP8","16h","18h"),
("03/12/2025","Communication cellulaire et signalisation TD2 GP7","16h","18h"),
("04/12/2025","Biochimie TD4 GP9","8h30","10h30"),
("04/12/2025","Communication cellulaire et signalisation TD2 GP10","8h30","10h30"),
("04/12/2025","Biochimie TD4 GP10","10h45","12h45"),
("04/12/2025","Communication cellulaire et signalisation TD2 GP9","10h45","12h45"),
("04/12/2025","Biochimie TD4 GP11","13h45","15h45"),
("04/12/2025","Communication cellulaire et signalisation TD2 GP12","13h45","15h45"),
("04/12/2025","Biochimie TD4 GP12","16h","18h"),
("04/12/2025","Communication cellulaire et signalisation TD2 GP11","16h","18h"),
]

SLOTS = ["8h30 - 10h30", "10h45 - 12h45", "13h45 - 15h45", "16h - 18h"]
SLOT_KEYS = [("8h30","10h30"), ("10h45","12h45"), ("13h45","15h45"), ("16h","18h")]

def normalize_gp(text: str) -> str:
    import re
    return re.sub(r'\\bG(\\d{1,2})\\b', r'GP\\1', text)

def build_gp_dict(rows):
    out = defaultdict(list)
    for d, label, start, end in rows:
        dt = datetime.strptime(d, "%d/%m/%Y")
        label = normalize_gp(label)
        import re
        m = re.search(r'\\bGP(\\d{1,2})\\b', label)
        if not m: 
            continue
        gp = f"GP{int(m.group(1))}"
        out[gp].append((dt, label, start.replace("H","h"), end.replace("H","h")))
    return out

def fr_week_start(d):
    return d - timedelta(days=d.weekday())

# ---------- Excel builder with openpyxl ONLY ----------
def make_excel(gp_courses: dict, start_range: datetime, end_range: datetime) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin', color='000000')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="E6F1FF", end_color="E6F1FF", fill_type="solid")
    slot_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")

    for gp_num in range(1, 13):
        gp = f"GP{gp_num}"
        ws = wb.create_sheet(title=gp)

        ws.append(["Semaine du :", "Heure", "Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"])
        for col in range(1, 10):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all

        cur = fr_week_start(start_range)
        while cur <= end_range:
            for si, slot in enumerate(SLOTS):
                row = [cur.strftime("%Y-%m-%d") if si==0 else None, slot] + [None]*7
                ws.append(row)
                r = ws.max_row
                for c in range(1, 10):
                    cell = ws.cell(row=r, column=c)
                    if c==2:
                        cell.fill = slot_fill
                    cell.alignment = Alignment(horizontal="center" if c==2 else "left", vertical="center")
                    cell.border = border_all

            day_cols = {0:3,1:4,2:5,3:6,4:7,5:8,6:9}
            base_row = ws.max_row-3
            for dt, label, start, end in gp_courses.get(gp, []):
                if fr_week_start(dt)==cur:
                    try:
                        slot_index = SLOT_KEYS.index((start, end))
                    except ValueError:
                        slot_index = 0
                    col = day_cols[dt.weekday()]
                    r = base_row + slot_index
                    cell = ws.cell(row=r, column=col)
                    content = f"{label} ({start}-{end})"
                    cell.value = (str(cell.value) + "\\n" + content) if cell.value else content
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            cur += timedelta(days=7)

        widths = [14, 14, 22, 22, 22, 22, 22, 22, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ---------- State & UI ----------
if "rows" not in st.session_state:
    st.session_state.rows = RAW.copy()

st.subheader("🔧 Ajouter / modifier un créneau")
col1, col2, col3, col4 = st.columns(4)
with col1:
    d = st.text_input("Date (jj/mm/aaaa)", value="24/09/2025")
with col2:
    start = st.text_input("Début (ex: 13h45)", value="13h45")
with col3:
    end = st.text_input("Fin (ex: 15h45)", value="15h45")
with col4:
    label = st.text_input("Intitulé (ex: Biochimie TD1 GP6)", value="Biochimie TD1 GP6")

c1, c2 = st.columns([1,1])
with c1:
    if st.button("➕ Ajouter"):
        st.session_state.rows.append((d, label, start, end))
with c2:
    if st.button("🗑️ Réinitialiser la liste aux données d'origine"):
        st.session_state.rows = RAW.copy()

st.divider()
st.subheader("📜 Aperçu des cours saisis")
preview = pd.DataFrame(st.session_state.rows, columns=["Date","Intitulé","Début","Fin"])
st.dataframe(preview, use_container_width=True, hide_index=True)

# Aperçu hebdo
st.divider()
st.subheader("👀 Aperçu hebdomadaire à l'écran")
gp_choice = st.selectbox("Groupe (GP)", [f"GP{i}" for i in range(1,13)], index=0)
week_start = st.date_input("Lundi de la semaine", value=datetime(2025,9,22))
week_start = datetime(week_start.year, week_start.month, week_start.day)

grid = pd.DataFrame({"Heure": SLOTS})
for day in ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]:
    grid[day] = ""

gp_dict = build_gp_dict(st.session_state.rows)
for dt, label, s, e in gp_dict.get(gp_choice, []):
    if fr_week_start(dt) == fr_week_start(week_start):
        try:
            idx = SLOT_KEYS.index((s,e))
        except ValueError:
            idx = 0
        day_name = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"][dt.weekday()]
        content = f"{label} ({s}-{e})"
        grid.loc[idx, day_name] = (grid.loc[idx, day_name] + "\\n" if grid.loc[idx, day_name] else "") + content

st.dataframe(grid, use_container_width=True, hide_index=True)

# Export
st.divider()
st.subheader("📤 Générer l'Excel par GP")
start_date = datetime(2025,9,22)
end_date = datetime(2025,12,4)
excel_bytes = make_excel(gp_dict, start_date, end_date)

st.download_button(
    "⬇️ Télécharger l'Excel (une feuille par GP)",
    data=excel_bytes,
    file_name="EDT_GP_hebdo_22sept-4dec.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.info("💡 Déploiement : repo GitHub avec `app.py` et `requirements.txt`, puis Streamlit Community Cloud.")

